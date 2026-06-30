"""Еженедельная сверка карточек товара (состав / свойства / способ применения)
из выгрузки 1С с Google-таблицей «ОП АС Фарм» → вкладка МАТРИЦА.

Логика:
1. Заходит в ящик Mail.ru по IMAP, берёт САМОЕ СВЕЖЕЕ письмо «Базовые рецептуры
   СТМ от DD.MM.YYYY» от Завода АС-ФАРМ (по дате в теме, не по дате получения),
   качает вложение xlsx.
2. Разбирает рецептуры (код = текст после последней запятой в «Наименование БР»).
3. По связке артикул→код 1С (MAP) сверяет три колонки МАТРИЦА: S(состав),
   T(свойства), U(способ применения).
4. Если ячейка пустая, а в 1С есть текст — заполняет (новинка, без подсветки).
   Если текст отличается (по смыслу, без учёта пробелов/переносов) — перезаписывает
   значением из 1С и подсвечивает ячейку красным с эскалацией оттенка
   (#f4cccc → #ea9999 → #e06666 → #cc0000). Владелец снимает заливку сам = сброс.
   Различия только в пробелах/переносах изменением НЕ считаются.
5. Формат строк НЕ трогает: пишет только значение + цвет заливки, при этом
   принудительно держит шрифт 7 и wrapStrategy=OVERFLOW (узкие строки). Высоту
   строк не меняет (это свойство строки, не ячейки).
6. При успешной сверке переносит письмо в Корзину. Если письма нет / нет вложения /
   сверка упала — письмо НЕ трогает.

Переменные окружения:
  MAILRU_USER       — логин ящика (по умолчанию nikol-oleinik@mail.ru)
  MAILRU_APP_PASS   — пароль приложения Mail.ru (обязателен)
  GSHEETS_SA_JSON   — содержимое JSON сервисного аккаунта (обязателен)
  DRY_RUN=1         — ничего не писать и письмо не удалять (только отчёт)
"""
from __future__ import annotations

import os
import re
import ssl
import json
import email
import imaplib
import tempfile
from email.header import decode_header

import openpyxl
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

SPREADSHEET_ID = "1sHlFGSVB-7V8V4q6kvcTR1rrw19EaabIaOgrCHU0DHE"
SHEET_NAME = "МАТРИЦА"
IMAP_HOST = "imap.mail.ru"
MAIL_USER = os.environ.get("MAILRU_USER", "nikol-oleinik@mail.ru")
MAIL_PASS = os.environ.get("MAILRU_APP_PASS", "")
DRY_RUN = os.environ.get("DRY_RUN") == "1"
SENDER = "as-farm-as-farm@yandex.ru"
SUBJECT_PREFIX = "базовые рецептуры стм"
TRASH = "&BBoEPgRABDcEOAQ9BDA-"  # «Корзина» в modified-UTF7

# колонки МАТРИЦА (0-based индексы): C=артикул(2), S=состав(18), T=свойства(19), U=способ(20)
COL_ART = 2
FIELDS = [("состав", 18, "sostav"), ("свойства", 19, "opisanie"), ("способ применения", 20, "sposob")]

# Связка: начало строки артикула (колонка C) -> код рецептуры 1С.
# Порядок важен: более длинные/специфичные ключи идут РАНЬШE коротких.
MAP = [
    ("Dental_40_natural", "АСП10.250318.01"),
    ("Dental_40_zemlyanika", "АСП10.250804.01"),
    ("Dental_100_banan", "АСП10.250805.01"),
    ("Dental_100_zemlyanika", "АСП10.250804.01"),
    ("Dental_20_zemlyanika", "АСП10.250804.01"),
    ("Dental_100", "АСП10.250721.01"),
    ("Dental_40", "АСП10.250721.01"),
    ("Dental20", "АСП10.250721.01"),
    ("Dental50", "АСП02.251121.01"),
    ("extract_romashka", "АС00.260513.2283"),
    ("extract_pihta", "АС00.260513.2284"),
    ("Zub_pasta_det", "АС39.251217.01"),
    ("Irrigator_500", "АС25.260507.01"),
    ("Irrigator_1000", "АС25.260507.01"),
    ("irrigator_new_1", "АС25.260507.01"),
    ("irrigator_new_05", "АС25.260507.01"),
    ("OralLubrikant", "АС00.260129.1020"),
    ("Oral_cherry", "АС00.260129.1274"),
    ("spray_hedonisme", "АС00.260130.1314"),
    ("OptikaSpray_new", "АС00/240312/13"),
    ("CrioGel1l", "АС00.240424.02"),
    ("CrioGel_5", "АС00.240424.02"),
    ("Crio_L25(new)", "АС21.250305.01"),
    ("CrioL50", "АС21.250305.01"),
    ("cryolipolysis25", "АС21.250305.01"),
    ("Cryolipolysis50", "АС21.250305.01"),
    ("энзимная черная мужская", "АС00.260227.1488"),
    ("энзимная арбуз", "АС00.260226.1504"),
]

# палитра эскалации (по нарастанию)
PALETTE = [(244, 204, 204), (234, 153, 153), (224, 102, 102), (204, 0, 0)]


def _dh(v):
    if not v:
        return ""
    return "".join(p.decode(e or "utf-8", "ignore") if isinstance(p, bytes) else p
                    for p, e in decode_header(v))


def norm(s):
    """Для сравнения: пробелы/переносы/регистр расхождением не считаем."""
    return re.sub(r"\s+", " ", (s or "")).strip().lower()


def code_for(cell):
    c = (cell or "").strip()
    for key, code in MAP:
        if c.startswith(key):
            return key, code
    return None, None


def imap_connect():
    M = imaplib.IMAP4_SSL(IMAP_HOST, ssl_context=ssl._create_unverified_context())
    M.login(MAIL_USER, MAIL_PASS)
    return M


def find_latest_recipe(M):
    """Возвращает (uid, subject, report_date) самого свежего письма-рецептуры или None."""
    M.select("INBOX")
    typ, data = M.uid("search", None, "FROM", SENDER)
    best = None
    for uid in data[0].split():
        typ, d = M.uid("fetch", uid, "(BODY.PEEK[HEADER.FIELDS (SUBJECT)])")
        subj = _dh(email.message_from_string(d[0][1].decode("utf-8", "ignore"))["Subject"])
        if not subj.lower().startswith(SUBJECT_PREFIX):
            continue
        m = re.search(r"от\s+(\d{2})\.(\d{2})\.(\d{4})", subj)
        key = (int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else (0, 0, 0)
        if best is None or key > best[2]:
            best = (uid.decode(), subj, key)
    return best


def download_xlsx(M, uid):
    typ, d = M.uid("fetch", uid, "(RFC822)")
    msg = email.message_from_bytes(d[0][1])
    for part in msg.walk():
        fn = part.get_filename()
        if fn and _dh(fn).lower().endswith((".xlsx", ".xls")):
            path = os.path.join(tempfile.gettempdir(), "recipes_1c.xlsx")
            with open(path, "wb") as f:
                f.write(part.get_payload(decode=True))
            return path
    return None


def parse_1c(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    data = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = r[1]
        if not name:
            continue
        code = str(name).split(",")[-1].strip().replace(" ", "")
        data[code] = {
            "sostav": (r[3] or "").strip(),
            "opisanie": (r[4] or "").strip(),
            "sposob": (r[5] or "").strip(),
        }
    return data


def color_level(bg):
    """Текущий уровень заливки: -1 нет/белая, 0..3 по палитре."""
    if not bg:
        return -1
    r = round(bg.get("red", 0) * 255)
    g = round(bg.get("green", 0) * 255)
    b = round(bg.get("blue", 0) * 255)
    for i, (R, G, B) in enumerate(PALETTE):
        if abs(r - R) <= 8 and abs(g - G) <= 8 and abs(b - B) <= 8:
            return i
    return -1  # белая или любая другая = считаем «снято»


def main():
    if not MAIL_PASS:
        raise SystemExit("Нет MAILRU_APP_PASS")
    sa_info = json.loads(os.environ["GSHEETS_SA_JSON"])
    creds = Credentials.from_service_account_info(
        sa_info, scopes=["https://www.googleapis.com/auth/spreadsheets"])
    svc = build("sheets", "v4", credentials=creds, cache_discovery=False)

    # sheetId вкладки МАТРИЦА
    meta = svc.spreadsheets().get(spreadsheetId=SPREADSHEET_ID,
                                  fields="sheets.properties").execute()
    sheet_id = next(s["properties"]["sheetId"] for s in meta["sheets"]
                    if s["properties"]["title"] == SHEET_NAME)

    print(f"Режим: {'DRY-RUN (ничего не меняю)' if DRY_RUN else 'БОЕВОЙ'}", flush=True)

    M = imap_connect()
    latest = find_latest_recipe(M)
    if not latest:
        print("Письмо-рецептура не найдено — выходим, ничего не делаю.", flush=True)
        M.logout()
        return
    uid, subj, rep = latest
    print(f"Свежее письмо: «{subj}»", flush=True)
    path = download_xlsx(M, uid)
    if not path:
        print("Во вложении нет xlsx — письмо НЕ трогаю, выходим.", flush=True)
        M.logout()
        return
    data = parse_1c(path)
    print(f"Рецептур в выгрузке: {len(data)}", flush=True)

    # значения и цвета колонок C..U
    rng = f"{SHEET_NAME}!A1:U200"
    grid = svc.spreadsheets().get(
        spreadsheetId=SPREADSHEET_ID, ranges=[rng], includeGridData=True,
        fields="sheets.data.rowData.values(formattedValue,userEnteredFormat.backgroundColor)"
    ).execute()
    rowdata = grid["sheets"][0]["data"][0].get("rowData", [])

    requests = []
    filled, changed, unchanged, problems = [], [], [], []
    for ridx, row in enumerate(rowdata):
        cells = row.get("values", [])
        art = cells[COL_ART].get("formattedValue", "") if len(cells) > COL_ART else ""
        if not art.strip():
            continue
        key, code = code_for(art)
        if not key:
            continue
        if code not in data:
            problems.append(f"стр.{ridx+1} {key}: кода {code} нет в выгрузке 1С")
            continue
        for label, col, dk in FIELDS:
            new = data[code][dk]
            cur = cells[col].get("formattedValue", "") if len(cells) > col else ""
            bg = cells[col].get("userEnteredFormat", {}).get("backgroundColor") if len(cells) > col else None
            if not new:
                continue
            if not cur.strip():
                # пустая ячейка → первичное заполнение, без подсветки
                filled.append(f"стр.{ridx+1} {key} / {label}")
                requests.append(_cell_req(sheet_id, ridx, col, new, level=None))
            elif norm(cur) != norm(new):
                lvl = color_level(bg)
                new_lvl = min((lvl if lvl >= 0 else -1) + 1, 3)
                changed.append(f"стр.{ridx+1} {key} / {label} (заливка ур.{new_lvl})")
                requests.append(_cell_req(sheet_id, ridx, col, new, level=new_lvl))
            else:
                unchanged.append(f"{key}/{label}")

    # отчёт
    print(f"\nСверка: совпало {len(unchanged)}, заполнено новых {len(filled)}, изменено {len(changed)}", flush=True)
    for x in filled:
        print("  [заполнено]", x, flush=True)
    for x in changed:
        print("  [ИЗМЕНЕНО]", x, flush=True)
    for x in problems:
        print("  [!]", x, flush=True)

    if not requests:
        print("\nИзменений нет.", flush=True)
    elif DRY_RUN:
        print(f"\nDRY-RUN: записал бы {len(requests)} ячеек, но ничего не трогаю.", flush=True)
    else:
        svc.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID, body={"requests": requests}).execute()
        print(f"\nЗаписано ячеек: {len(requests)}.", flush=True)

    # перенос письма в Корзину — только в боевом режиме и без проблем
    if DRY_RUN:
        print("DRY-RUN: письмо оставляю на месте.", flush=True)
    elif problems:
        print("Есть нерешённые связки — письмо НЕ удаляю, оставляю на месте.", flush=True)
    else:
        M.select("INBOX")
        M.uid("copy", uid, TRASH)
        M.uid("store", uid, "+FLAGS", "\\Deleted")
        M.expunge()
        print("Письмо перенесено в Корзину.", flush=True)
    M.logout()


def _cell_req(sheet_id, ridx, col, value, level):
    """updateCells: значение + (опц.) цвет заливки + шрифт 7 + OVERFLOW.
    Прочие свойства ячейки (границы, выравнивание) и высота строки не трогаются."""
    fmt = {
        "textFormat": {"fontSize": 7},
        "wrapStrategy": "OVERFLOW",
    }
    mask = "userEnteredValue,userEnteredFormat.textFormat.fontSize,userEnteredFormat.wrapStrategy"
    if level is not None:
        R, G, B = PALETTE[level]
        fmt["backgroundColor"] = {"red": R / 255, "green": G / 255, "blue": B / 255}
        mask += ",userEnteredFormat.backgroundColor"
    return {
        "updateCells": {
            "rows": [{"values": [{
                "userEnteredValue": {"stringValue": value},
                "userEnteredFormat": fmt,
            }]}],
            "fields": mask,
            "start": {"sheetId": sheet_id, "rowIndex": ridx, "columnIndex": col},
        }
    }


if __name__ == "__main__":
    main()
